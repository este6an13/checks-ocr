import os
import shutil
import pandas as pd
from openpyxl.styles import PatternFill

from constants import CONFIDENCE_COLUMNS
from utils import get_id


def save_image(img, id, images_folder):
    """
    Save PIL image as PNG with a unique filename generated using uuid in the 'images' folder.

    :param img: The PIL image object.
    :param images_folder: The name of the folder to save images (default is 'images').
    :return: The filename (including path) of the saved PNG file.
    """
    try:
        # Create the images folder if it doesn't exist
        os.makedirs(images_folder, exist_ok=True)

        # Generate a unique filename using uuid
        filename = os.path.join(images_folder, id + ".png")

        # Save the image as PNG
        img.save(filename, format="PNG")

        return filename
    except Exception as e:
        print(f"Error saving image: {e}")
        return None


def load_data(file_path="../../data.xlsx", sheet_name=0, columns=None):
    """
    Read an Excel file and load it into a pandas DataFrame.
    If the file doesn't exist, it will be created.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name or index of the sheet to read (default is 0).
    :param columns: List of column names to select from the Excel file (default is None, which reads all columns).
    :return: A pandas DataFrame containing the data.
    """
    try:
        # Check if the file exists, create it if not
        if not os.path.exists(file_path):
            # Create the Excel file with an empty DataFrame if it doesn't exist
            pd.DataFrame(columns=columns).to_excel(file_path, index=False)

        # Read the Excel file into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=columns)
        return df
    except Exception as e:
        print(f"Error: Unable to read Excel file. {e}")
        return None


def get_unprocessed_filenames(folder_path):
    """
    Get the names of all PDF files in a specified folder.

    :param folder_path: Path to the folder (default is 'unprocessed').
    :return: A list of PDF file names.
    """
    pdf_files = []
    try:
        folder_path = os.path.abspath(folder_path)
        for file_name in os.listdir(folder_path):
            if file_name.lower().endswith(".pdf"):
                pdf_files.append(file_name)
        return pdf_files
    except FileNotFoundError:
        print(f"Error: Folder '{folder_path}' not found.")
        return None
    except Exception as e:
        print(f"Error: Unable to retrieve PDF files. {e}")
        return None


def check_if_processed(df, filename):
    """
    Check if a filename (without the .pdf extension) is in the 'ID' column of the DataFrame.

    :param df: The DataFrame containing the 'ID' column.
    :param filename: The filename to check.
    :return: True if the filename is in the 'ID' column, False otherwise.
    """
    try:
        # Extract the filename without the extension
        id = get_id(filename)

        # Check if the filename is in the 'ID' column
        return id in df["ID"].values
    except Exception as e:
        print(f"Error: Unable to check if processed. {e}")
        return False


def move_file(filename, id, source_folder, destination_folder):
    """
    Move a file from the source folder to the destination folder.

    :param filename: The name of the file to move.
    :param source_folder: The name of the source folder (default is 'unprocessed').
    :param destination_folder: The name of the destination folder (default is 'processed').
    :return: True if the move is successful, False otherwise.
    """
    try:
        # Create the destination folder if it doesn't exist
        os.makedirs(destination_folder, exist_ok=True)

        # Construct paths for source and destination
        source_path = os.path.join(source_folder, filename)
        destination_path = os.path.join(destination_folder, f"{id}.pdf")

        # Move the file
        shutil.move(source_path, destination_path)

        return True
    except Exception as e:
        print(f"Error moving file: {e}")
        return False


def write_confidence(df, filename, sheet_name="data"):
    """
    Write a DataFrame to an Excel file.

    :param df: The DataFrame to write.
    :param filename: The name of the Excel file (default is 'confidence.xlsx').
    :param sheet_name: The name of the sheet in the Excel file (default is 'data').
    :return: True if the write is successful, False otherwise.
    """
    try:
        # Write the DataFrame to Excel
        df.to_excel(filename, sheet_name=sheet_name, index=False)

        return True
    except Exception as e:
        print(f"Error writing DataFrame to Excel: {e}")
        return False


def write_data(df, confidence_df, data_path, images_path):
    # Create a new Excel workbook and add the DataFrame to it
    with pd.ExcelWriter(data_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="data", index=False)

        # Access the worksheet
        wb = writer.book
        ws = wb["data"]

        # Define fill styles for conditional formatting
        colors = ["b0ffb1", "ddffb0", "fffab0", "f0bca8", "ffffff", "e9d7fc"]
        fill_0 = PatternFill(
            start_color=colors[0], end_color=colors[0], fill_type="solid"
        )
        fill_1 = PatternFill(
            start_color=colors[1], end_color=colors[1], fill_type="solid"
        )
        fill_2 = PatternFill(
            start_color=colors[2], end_color=colors[2], fill_type="solid"
        )
        fill_3 = PatternFill(
            start_color=colors[3], end_color=colors[3], fill_type="solid"
        )
        fill_4 = PatternFill(
            start_color=colors[4], end_color=colors[4], fill_type="solid"
        )
        fill_5 = PatternFill(
            start_color=colors[5], end_color=colors[5], fill_type="solid"
        )

        # Apply conditional formatting based on confidence_df values
        for col in CONFIDENCE_COLUMNS:
            col_index = df.columns.get_loc(
                col
            )  # Get column index (starting from 1)
            for i, cell in enumerate(
                ws["{}:{}".format(chr(65 + col_index), chr(65 + col_index))]
            ):
                if i == 0:  # Skip the header row
                    continue
                confidence_value = confidence_df.iloc[i - 1][col]
                if confidence_value > 98:
                    cell.fill = fill_0
                elif confidence_value > 95:
                    cell.fill = fill_1
                elif confidence_value > 90:
                    cell.fill = fill_2
                elif confidence_value > 0:
                    cell.fill = fill_3
                elif confidence_value == 0:  # not used
                    cell.fill = fill_4
                else:
                    cell.fill = fill_5

        # Add hyperlinks to the 'ID' column
        id_col_index = df.columns.get_loc("ID") + 1
        for row in ws.iter_rows(
            min_col=id_col_index, max_col=id_col_index, min_row=2
        ):
            cell = row[0]
            image_path = f"{images_path}/{cell.value}.png"
            cell.hyperlink = image_path
